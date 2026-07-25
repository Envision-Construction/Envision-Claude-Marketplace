#!/usr/bin/env python3
"""
render_xlsx.py — render extracted JSON to an EDGAR-styled xlsx workbook.

This is the highest-priority renderer because financial statements arrive most
often as xlsx and users expect xlsx out. It implements the segment-mode visual
spec from references/financial-statements.md:

  - Times New Roman 10pt body, 11pt for the registrant name, 10pt bold ALL CAPS
    statement title
  - Title block centered above the data, three lines
  - Period header right-aligned over numeric columns
  - Year column headers right-aligned + bold, single rule underneath
  - Line labels left-aligned, indented per row.indent (0.25" / 3 chars per level)
  - Numbers right-aligned, accounting format with thousands separators
  - Negatives in parentheses, NOT minus signs
  - Explicit `$` only on (a) the first data row and (b) subtotal / grand-total
    rows — implemented via the "\"$\"#,##0_);(\"$\"#,##0);\"—\";@" accounting
    format, with a "no-$" variant for intermediate rows. We deliberately do
    NOT use Excel's `_($* #,##0_)` floating-$ asterisk-fill token because
    Google Sheets silently drops it on import — see the note above the FMT_*
    constants. Both Excel and Sheets render the current formats identically.
  - Em dash for zero values
  - Subtotal row: thin top border on numeric cells
  - Grand total row: thin top border + double bottom border on numeric cells
  - No bold in body — emphasis is rules, not weight
  - Multi-segment workbooks get one sheet per segment, sheet name = short
    statement code (IS, BS, CF, SE, CI)

Usage:
    python3 render_xlsx.py <extracted.json> <output.xlsx>
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required: {exc}")


_SHEET_CODE = {
    "income_statement": "IS",
    "balance_sheet": "BS",
    "cash_flows": "CF",
    "stockholders_equity": "SE",
    "comprehensive_income": "CI",
}

# Accounting formats — explicit "$", parens for negatives, em dash for zero,
# right-edge alignment of positives vs parenthesized negatives via the trailing
# `_)` token (reserves one paren width on the positive branch).
#
# We do NOT use Excel's classic floating-$ accounting format
# (`_($* #,##0_);_($* (#,##0);_($* "—"_);_(@_)`) because the `* ` fill-repeat
# token is silently dropped by Google Sheets on xlsx import — the $ glues to
# the digits and column alignment breaks. The formats below render identically
# in Excel desktop and Google Sheets. Tradeoff: we lose the floating-$ visual
# that classic EDGAR filings have. The skill prioritizes "looks correct
# everywhere" over "looks pixel-perfect EDGAR in Excel only".
FMT_DOLLAR    = '"$"#,##0_);("$"#,##0);"—";@'
FMT_NUMBER    = '#,##0_);(#,##0);"—";@'
FMT_EPS       = '"$"#,##0.00_);("$"#,##0.00);"—";@'  # for per-share
FMT_SHARES    = '#,##0_);(#,##0);"—";@'              # share counts


def _font(size: int = 10, bold: bool = False, italic: bool = False) -> Font:
    return Font(name="Times New Roman", size=size, bold=bold, italic=italic, color="000000")


def _border_top_single() -> Border:
    return Border(top=Side(style="thin", color="000000"))


def _border_top_single_bottom_double() -> Border:
    return Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )


def _is_eps_row(label: str) -> bool:
    l = label.lower()
    return "per share" in l or l.strip() in {"basic", "diluted"}


def _is_share_count_row(label: str) -> bool:
    l = label.lower()
    return "weighted-average" in l or "weighted average" in l or "shares outstanding" in l


def _indent_prefix(indent: int) -> str:
    return "   " * indent  # 3 spaces per level — keeps Excel column widths sane


def _render_segment(wb: Workbook, segment: dict[str, Any], metadata: dict[str, Any]) -> None:
    stype = segment["statement_type"]
    code = _SHEET_CODE.get(stype, "STMT")
    sheet_name = code
    suffix = 2
    while sheet_name in wb.sheetnames:
        sheet_name = f"{code}_{suffix}"
        suffix += 1
    ws = wb.create_sheet(title=sheet_name)

    period_labels = segment.get("period_labels", []) or []
    n_periods = max(len(period_labels), 1)
    # Column layout: A = label, B = optional dollar-sign anchor (handled by
    # accounting format), C..(C + n_periods - 1) = period columns.
    label_col = 1
    first_num_col = 2
    last_num_col = first_num_col + n_periods - 1

    # Column widths — label wide enough for indented sub-line items
    ws.column_dimensions[get_column_letter(label_col)].width = 56
    for i in range(first_num_col, last_num_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Margins + page setup — Letter, 1" margins
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_options.horizontalCentered = True

    row = 1

    # --- Title block (3 lines centered across label + period cols) -----------
    registrant = (metadata.get("registrant") or "").strip()
    if registrant:
        ws.cell(row=row, column=label_col, value=registrant)
        ws.cell(row=row, column=label_col).font = _font(11, bold=True)
        ws.cell(row=row, column=label_col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=label_col, end_row=row, end_column=last_num_col)
        row += 1

    title = segment.get("title", "")
    ws.cell(row=row, column=label_col, value=title)
    ws.cell(row=row, column=label_col).font = _font(10, bold=True)
    ws.cell(row=row, column=label_col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=label_col, end_row=row, end_column=last_num_col)
    row += 1

    units = segment.get("units_note", "")
    if units:
        ws.cell(row=row, column=label_col, value=units)
        ws.cell(row=row, column=label_col).font = _font(9, italic=True)
        ws.cell(row=row, column=label_col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=label_col, end_row=row, end_column=last_num_col)
        row += 1

    row += 1  # spacer

    # --- Period header (e.g., "Years Ended December 31,") -------------------
    period_header = segment.get("period_header", "")
    if period_header:
        # span across the numeric columns, right-aligned via the last column
        ws.cell(row=row, column=first_num_col, value=period_header)
        ws.merge_cells(
            start_row=row, start_column=first_num_col,
            end_row=row, end_column=last_num_col,
        )
        ws.cell(row=row, column=first_num_col).font = _font(10)
        ws.cell(row=row, column=first_num_col).alignment = Alignment(horizontal="center")
        row += 1

    # --- Year/date column headers (bold, right-aligned, single rule under) ---
    for i, label in enumerate(period_labels):
        cell = ws.cell(row=row, column=first_num_col + i, value=label)
        cell.font = _font(10, bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(bottom=Side(style="thin", color="000000"))
    # If no period labels, still write a rule via a blank-row underline
    if not period_labels:
        for i in range(first_num_col, last_num_col + 1):
            ws.cell(row=row, column=i, value="").border = Border(
                bottom=Side(style="thin", color="000000")
            )
    row += 1

    # --- Body rows ----------------------------------------------------------
    first_data_row_written = False
    current_section = ""  # tracks most recent section_header for EPS/shares context
    for r in segment.get("rows", []):
        label = r.get("label", "")
        role = r.get("role", "data")
        indent = int(r.get("indent", 0) or 0)
        values = r.get("values", [])

        if role == "spacer":
            row += 1
            continue

        # Label cell
        lc = ws.cell(row=row, column=label_col,
                     value=_indent_prefix(indent) + (label or ""))
        lc.font = _font(10)
        lc.alignment = Alignment(horizontal="left", vertical="top")

        if role in {"banner"}:
            # ALL CAPS section banner — ASSETS / LIABILITIES AND...
            lc.value = (label or "").upper()
            lc.font = _font(10)  # regular weight per EDGAR convention
            row += 1
            continue
        if role == "section_header":
            # title-case subsection header with trailing colon — already in
            # label; no numeric values. Track it so we can disambiguate
            # "Basic"/"Diluted" rows that follow.
            current_section = (label or "").lower()
            row += 1
            continue
        if role == "text_only":
            # e.g., "Commitments and contingencies"
            row += 1
            continue

        # Numeric row — decide format. Disambiguate "Basic"/"Diluted" rows
        # using the current_section context: under "Earnings per ..." use EPS
        # format; under "Weighted-average ... shares outstanding:" use share
        # count integer format.
        section_is_shares = "shares outstanding" in current_section or "weighted-average" in current_section
        section_is_eps = "per share" in current_section or "earnings per" in current_section
        if section_is_shares or _is_share_count_row(label):
            fmt_first = FMT_SHARES
            fmt_rest = FMT_SHARES
        elif section_is_eps or _is_eps_row(label):
            fmt_first = FMT_EPS
            fmt_rest = '#,##0.00_);(#,##0.00);"—";@'
        else:
            fmt_first = FMT_DOLLAR
            fmt_rest = FMT_NUMBER

        # Float $ on:
        #   - the first data row of the segment
        #   - every subtotal and grand_total row
        floats_dollar = (
            role in {"subtotal", "grand_total"} or not first_data_row_written
        )

        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=first_num_col + i)
            cell.font = _font(10)
            cell.alignment = Alignment(horizontal="right")
            if v is None:
                cell.value = "—"
            else:
                cell.value = float(v)
                cell.number_format = fmt_first if (floats_dollar and i < n_periods) else fmt_rest

        if role == "subtotal":
            for i in range(n_periods):
                ws.cell(row=row, column=first_num_col + i).border = _border_top_single()
        elif role == "grand_total":
            for i in range(n_periods):
                ws.cell(row=row, column=first_num_col + i).border = _border_top_single_bottom_double()

        if role == "data" and any(v is not None for v in values):
            first_data_row_written = True
        row += 1

    # Footer page number anchor — centered "F-1" style
    ws.oddFooter.center.text = "F-1"
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "Times New Roman"


def render(extracted: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()
    # Drop the default sheet — we'll create per-segment sheets below
    default = wb.active
    wb.remove(default)

    metadata = extracted.get("metadata") or {}
    segments = extracted.get("segments") or []
    if not segments:
        # Fall back to a placeholder sheet that explains the situation —
        # full-filing mode is intentionally not implemented for xlsx because a
        # narrative 10-K does not belong in a spreadsheet. Tell the user.
        ws = wb.create_sheet(title="README")
        ws.cell(row=1, column=1, value=(
            "edgar-format render_xlsx.py: no statement segments detected.\n\n"
            "xlsx output is only emitted for financial-statement segments "
            "(income statement, balance sheet, cash flows, stockholders' "
            "equity, comprehensive income).\n\n"
            "For full-filing output (cover page + narrative + statements), "
            "re-run with render_html.py, render_pdf.py, or render_docx.py."
        ))
        ws.cell(row=1, column=1).font = _font(10)
        ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 90
    else:
        for seg in segments:
            _render_segment(wb, seg, metadata)

    wb.save(out_path)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("extracted", type=Path, help="JSON from extract.py")
    ap.add_argument("output", type=Path, help="output xlsx path")
    args = ap.parse_args(argv)

    data = json.loads(args.extracted.read_text(encoding="utf-8"))
    render(data, args.output)
    print(f"wrote {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
